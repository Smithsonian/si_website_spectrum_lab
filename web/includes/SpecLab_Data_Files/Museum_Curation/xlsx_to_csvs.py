import sys
from os.path import extsep, splitext
from openpyxl import load_workbook

filepath = sys.argv[1]

wb = load_workbook(filepath)

def is_value(v):
    return v is not None and v != "None"

for name in wb.sheetnames:
    ws = wb[name]

    # We'll assume that the first column is wavelength
    # and that the column range is contiguous
    # and that we have a header
    columns = ws.columns
    wavelength_column = next(columns)
    wavelengths = [str(cell.value) for cell in wavelength_column if cell.value is not None]

    # Convert nm -> microns
    wavelengths = [wavelengths[0]] + [float(w) / 1000 for w in wavelengths[1:]]

    for column in columns:
        values = [str(cell.value) for cell in column]
        items = []
        for wl, v in zip(wavelengths, values):
            if is_value(wl) and is_value(v):
                items.append(f"{wl},{v}")
        text = "\n".join(items)
        name, _ext = splitext(filepath)
        name = name.replace(" ", "_")
        colname = values[0].rstrip().replace(" ", "_")
        csv_path = f"{name}_{colname}{extsep}csv"
        with open(csv_path, 'w') as f:
            f.write(text)
